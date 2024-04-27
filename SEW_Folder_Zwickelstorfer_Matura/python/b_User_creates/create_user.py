import csv
import logging
import math
import random
import re
import os
import argparse
from logging.handlers import RotatingFileHandler

import unicodedata
from openpyxl import load_workbook

__author__ = "Zwickelstorfer Felix"

from openpyxl.workbook import Workbook


def setUpLogger(path: str) -> None:
    """ creates the logger """
    global logger
    logger = logging.getLogger('my_logger')
    logger.setLevel(logging.INFO)

    os.makedirs(path, exist_ok=True)
    file_handler = RotatingFileHandler(getSafeFilePaths(path)[3], maxBytes=10000, backupCount=5)
    file_handler.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s %(message)s'))
    stream_handler = logging.StreamHandler()
    stream_formatter = logging.Formatter('%(levelname)s - %(message)s')
    stream_handler.setFormatter(stream_formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    logger.info("Logging turned on " + str(logger.level))


def start_program() -> None:
    """runs the logic for the program"""

    parser = argparse.ArgumentParser()

    parser.add_argument("-q", "--quiet", help='Only Log Errors', required=False, action='store_true')
    parser.add_argument("-v", "--verbose", help='Log Verbose Debug', required=False, action='store_true')

    args = parser.parse_args()
    global logger
    if "logger" not in globals():
        setUpLogger("./outUser")
    if args.verbose:
        logger.setLevel(logging.DEBUG)
    elif args.quiet:
        logger.setLevel(logging.ERROR)
    else:
        logger.setLevel(logging.INFO)

    create_user("./res/Namen.xlsx", "./outUser")


def getSafeFilePaths(outputDir: str) -> tuple[str, str, str, str]:
    """returns the names for the files to safe"""
    return f"{outputDir}/creates.sh", f"{outputDir}/deletes.sh", f"{outputDir}/list.csv", f"{outputDir}/logfile.log"


def getUserNameSpecialChars(username: str) -> str:
    """
    deletes/replaces all special chars and removes them if necessary
    :param username: the original name
    >>> print(getUserNameSpecialChars("1Hallo"))
    k1hallo
    >>> print(getUserNameSpecialChars("Hällo"))
    haello
    >>> print(getUserNameSpecialChars("Hàllo"))
    hallo
    >>> print(getUserNameSpecialChars("Hàllo"))
    hallo
    >>> print(getUserNameSpecialChars("Hàlloßen"))
    hallossen
    >>> print(getUserNameSpecialChars("Hàllo Welt"))
    hallo_welt
    >>> print(getUserNameSpecialChars("Hàllo Weltپ!پ"))
    hallo_welt
    """
    uName = str(username).lower()
    if uName[0].isnumeric():
        uName = f"k{uName}"
    uName = uName.replace("ä", "ae")
    uName = uName.replace("ö", "oe")
    uName = uName.replace("ü", "ue")
    uName = uName.replace("ß", "ss")
    norm_txt = unicodedata.normalize('NFD', uName)
    shaved = ''.join(c for c in norm_txt if not unicodedata.combining(c))
    uName = unicodedata.normalize('NFC', shaved)
    uName = re.sub("[^0-9a-z ]", "", uName)
    uName = uName.replace(" ", "_")
    return uName


def addCreateCommand(createFile, x: tuple[str, str, str, str], addedUsers: list[str], listingFile) -> str:
    """add the command line to create a class user"""
    uName = getUserNameSpecialChars(x[1])
    startUName = uName
    counter = 1
    while uName in addedUsers:
        uName = startUName + str(counter)
        counter += 1
    uName = getUserNameSpecialChars(uName)
    addedUsers.append(uName)
    createFile.write(
        f'useradd -d "/home/klassen/{uName}" -c "{x[0]} - {x[1]}" -m -g "{x[2]}" -G cdrom,plugdev,sambashare -s /bin/bash {uName}\n')
    passwd = f"{str(x[0]).lower()}{get_password_random_char()}{x[1]}{get_password_random_char()}{x[2].lower()}{get_password_random_char()}"
    createFile.write(f'echo "{uName}:{passwd}" | chpasswd\n')
    listingFile.write(f'"{uName}","{passwd}"\n')
    return uName


def get_password_random_char() -> None:
    """returns a random char for the password"""
    passwdRndChar = "!%&(),._-=^#5"
    return passwdRndChar[math.floor(random.random() * len(passwdRndChar))]


def addDeleteCommand(deleteFile, uName: str) -> None:
    """add the command line to delete a class user"""
    deleteFile.write("userdel -r " + str(uName) + "\n")


def addCommands(createFile, deleteFile, listingFile, x: tuple[str, str, str, str], addedUsers: list[str]):
    """calls the commands to add the user"""
    uName = addCreateCommand(createFile, x, addedUsers, listingFile)
    addDeleteCommand(deleteFile, uName)
    logger.debug(f"Added class-user {x[0]} sucessfully")


def create_user(path: str, outputDir: str) -> None:
    """
    creates files to create and delete classes
    :param path: the path to the xlsx file
    :param outputDir:  the directory to put the files into
    >>> create_user("./res/Namen.xlsx", "./outUser")
    """
    global logger
    if "logger" not in globals():
        setUpLogger(outputDir)
    try:
        os.makedirs(outputDir, exist_ok=True)
        wb = load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        createPath, deletePath, listingPath, logPath = getSafeFilePaths(outputDir)
        with open(createPath, "w") as createFile, open(deletePath, "w") as deleteFile, open(listingPath, "w") as listingFile:
            createFile.write("#! /bin/sh\ngroupadd student\n")
            createFile.write("groupadd teacher\n")
            deleteFile.write("#! /bin/sh\n")
            listingFile.write("Username,Password\n")

            addedUsers = []
            for row in ws.iter_rows(min_row=2):
                x = row[0].value, row[1].value, row[2].value, row[3].value
                if x[0] is None:
                    break
                addCommands(createFile, deleteFile, listingFile, x, addedUsers)

        logger.info("Added classes successfully.")

    except FileNotFoundError:
        logger.critical(f"Couldn't find file {path}")
    except Exception:
        logger.critical(f"Couldn't access file {path}")
    to_excel(getSafeFilePaths(outputDir)[2],outputDir+"/list.xlsx")


def to_excel(csv_file, excel_file):
    # Create a new Excel workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    # Open the CSV file and read its content
    with open(csv_file, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        # Iterate over the rows in the CSV file and write them to the Excel sheet
        for row_index, row in enumerate(csvreader, start=1):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

    # Save the Excel file
    workbook.save(excel_file)
if __name__ == "__main__":
    start_program()
