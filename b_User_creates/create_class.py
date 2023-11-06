import logging
import math
import random
import re
import sys
import os
import argparse
from logging.handlers import RotatingFileHandler

import unicodedata
from openpyxl import load_workbook

__author__ = "Zwickelstorfer Felix"


def setUpLogger() -> None:
    """ creates the logger """
    global logger
    logger = logging.getLogger('my_logger')
    logger.setLevel(logging.INFO)
    file_handler = RotatingFileHandler('classes_logfile.log', maxBytes=10000, backupCount=5)
    file_handler.setFormatter(logging.Formatter('[%(asctime)s] %(levelname)s %(message)s'))
    stream_handler = logging.StreamHandler()
    stream_formatter = logging.Formatter('%(levelname)s - %(message)s')
    stream_handler.setFormatter(stream_formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    logger.info("Logging turned on " + str(logger.level))


def start_program() -> None:
    """runs the logic for the program"""

    parser = argparse.ArgumentParser()

    parser.add_argument("-q", "--quiet", help='Only Log Errors', required=False, action='store_true')
    parser.add_argument("-v", "--verbose", help='Log Verbose Debug', required=False, action='store_true')

    args = parser.parse_args()
    global logger
    if "logger" not in globals():
        setUpLogger()
    if args.verbose:
        logger.setLevel(logging.DEBUG)
    elif args.quiet:
        logger.setLevel(logging.ERROR)
    else:
        logger.setLevel(logging.INFO)

    createClasses("./res/Klassenraeume_2023.xlsx", "./outClasses")


#  # TODO change
#     logging.basicConfig(level=getattr(logging, args), format='[%(asctime)s] %(levelname)s %(message)s',
#                             datefmt='%Y-%m-%d %H:%M:%S')
#     logging.log(logging.INFO, "Logging turned on " + args.loglevel)
#     if args.keygen:
#         minSize = 17
#         if args.keygen < minSize:
#             logging.log(logging.ERROR, "Keygen bit size must be greater than " + str(minSize))
#         else:
#             logging.log(logging.INFO, "Trying to generate keys with " + str(args.keygen) + " bit")


def getSafeFilePaths(outputDir: str) -> tuple[str, str, str]:
    """returns the names for the files to safe"""
    return f"{outputDir}/creates.sh", f"{outputDir}/deletes.sh", f"{outputDir}/list.csv", f"{outputDir}/logfile.log"


def getUserNameSpecialChars(username: str) -> str:
    """
    deletes/replaces all special chars and removes them if necessary
    :param username: the original name
    >>> print(getUserNameSpecialChars("1Hallo"))
    k1hallo
    >>> print(getUserNameSpecialChars("Hällo"))
    haello
    >>> print(getUserNameSpecialChars("Hàllo"))
    hallo
    >>> print(getUserNameSpecialChars("Hàllo"))
    hallo
    >>> print(getUserNameSpecialChars("Hàlloßen"))
    hallossen
    >>> print(getUserNameSpecialChars("Hàllo Welt"))
    hallo_welt
    >>> print(getUserNameSpecialChars("Hàllo Weltپ!پ"))
    hallo_welt
    """
    uName = str(username).lower()
    if uName[0].isnumeric():
        uName = f"k{uName}"
    uName = uName.replace("ä", "ae")
    uName = uName.replace("ö", "oe")
    uName = uName.replace("ü", "ue")
    uName = uName.replace("ß", "ss")
    norm_txt = unicodedata.normalize('NFD', uName)
    shaved = ''.join(c for c in norm_txt if not unicodedata.combining(c))
    uName = unicodedata.normalize('NFC', shaved)
    uName = re.sub("[^0-9a-z ]", "", uName)
    uName = uName.replace(" ", "_")
    return uName


def addCreateCommand(createFile, x: tuple[str, str, str], addedUsers: list[str]) -> None:
    """add the command line to create a class user"""
    uName = getUserNameSpecialChars(x[0])
    if uName in addedUsers:
        logger.critical(f"User ${uName} already exists!")
        sys.exit(1)
    addedUsers.append(uName)
    createFile.write(
        f'useradd -d "/home/klassen/{uName}" -c "{x[0]} - {x[2]}" -m -g "klasse" -G cdrom,plugdev,sambashare -s /bin/bash {uName}\n')
    createFile.write(
        f'echo "{uName}:{str(x[0]).lower()}{get_password_random_char()}{x[1]}{get_password_random_char()}{x[2].lower()}{get_password_random_char()}" | chpasswd\n')


def get_password_random_char() -> None:
    """returns a random char for the password"""
    passwdRndChar = "!%&(),._-=^#5"
    return passwdRndChar[math.floor(random.random() * len(passwdRndChar))]


def addDeleteCommand(deleteFile, x: tuple[str, str, str]) -> None:
    """add the command line to delete a class user"""
    uName = getUserNameSpecialChars(x[0])
    deleteFile.write("userdel -r " + str(uName) + "\n")


def addCommands(createFile, deleteFile, x: tuple[str, str, str], addedUsers: list[str]):
    """calls the commands to add the user"""
    addCreateCommand(createFile, x, addedUsers)
    addDeleteCommand(deleteFile, x)
    logger.debug(f"Added class-user {x[0]} sucessfully")


def createClasses(path: str, outputDir: str) -> None:
    """
    creates files to create and delete classes
    :param path: the path to the xlsx file
    :param outputDir:  the directory to put the files into
    >>> createClasses("./res/Klassenraeume_2023.xlsx", "./outClasses")
    >>> createClasses("./res/Klassenraeume_2023_With_Double.xlsx", "./outClasses")
    Traceback (most recent call last):
        ...
    SystemExit: 1
    >>> createClasses("./res/Klassenraeume_2023_NOT_EXISTING.xlsx", "./outClasses")
    """
    global logger
    if "logger" not in globals():
        setUpLogger()
    try:
        os.makedirs(outputDir, exist_ok=True)
        wb = load_workbook(path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        createPath, deletePath, listingPath, logPath = getSafeFilePaths(outputDir)
        with open(createPath, "w") as createFile, open(deletePath, "w") as deleteFile:
            createFile.write("#! /bin/sh\ngroupadd klasse\n")
            deleteFile.write("#! /bin/sh\n")

            addedUsers = []
            for row in ws.iter_rows(min_row=2):
                x = row[0].value, row[1].value, row[2].value
                if x[0] is None:
                    break
                addCommands(createFile, deleteFile, x, addedUsers)

            addCommands(createFile, deleteFile, ("lehrer", "0", "JUE"), addedUsers)  # Jüngling
            addCommands(createFile, deleteFile, ("seminar", "0", "JUE"), addedUsers)  # Jüngling

        logger.info("Added classes successfully.")
    except FileNotFoundError:
        logger.critical(f"Couldn't find file {path}")
    except Exception:
        logger.critical(f"Couldn't access file {path}")


# - ein Bash-Script2 mit allen notwendigen Schritten/Befehlen zum Erzeugen der Benutzer3
# – ein Bash-Script zum Löschen der angelegten Benutzer
# – eine Liste mit Usernamen und Passwörtern zum Verteilen an die unterrichtenden Lehrer
# – ein Logfile mit sinnvollen Angaben


# – der Dateiname für die Eingabedatei soll auf der Kommandozeile angegeben werden (ohne vorangestelltes
# -f o.ä.).
# – mit den Optionen -v für verbose und -q für quite soll der Umfang des Loggings eingestellt werden können
# (Loglevel).
# ∗ damit sollten sich auch die Bash-Scripts ändern

# Die Home-Verzeichnisse der Benutzer soll unter /home/klassen bzw. /home/lehrer/ angelegt werden,

# – Wichtig: eventuell braucht man im Script ein Escape (“\”) beim Passwort, aber nicht im Textfile!

# – bei bereits existierenden Benutzern sollte eine Fehlermeldung erfolgen.
# – die Bash-Scripts sollten bei Problemen abbrechen

# – Logging
# ∗ Wir wollen einen RotatingFileHandler mit maximal 10.000 Bytes7 und maximal 5 alte Versionen
# (backup).
# ∗ Zusätzlich ist ein StreamHandler (=Ausgabe auf der Konsole) anzulegen.
# ∗ Tipp: je nach Kommandozeilen-Parametern kann man zB. logging.basicConfig(level=logging.DEBUG)
# oder logger.setLevel(logging.WARNING) aufrufen.
# ∗ Alternative: logging.config.fileConfig('logging.conf')
# ∗ es sollte für jede Art von Meldung mindestens einen Aufruf geben
# · Fehler “Datei nicht gefunden” statt Exception
# · Debug: alle Details zur Fehlersuche


if __name__ == "__main__":
    start_program()
